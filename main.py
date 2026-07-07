"""
LORANNYS LUCAS OS — Render API
================================
Backend mínimo para las únicas dos cosas que Supabase no puede hacer:
  1. Rule Engine Python (análisis de transcripciones en español)
  2. Generación de archivos Excel y PDF

Deploy: Render.com → New Web Service → Free tier
  Build Command : pip install -r requirements.txt
  Start Command : uvicorn main:app --host 0.0.0.0 --port $PORT

Variables de entorno (opcional):
  ALLOWED_ORIGINS  — dominios del frontend separados por coma
                     ej: https://mi-dominio.com,https://lorannys.vercel.app
                     Si no se define, se permite cualquier origen (modo desarrollo)
"""

import io
import os
import unicodedata
from datetime import date, timedelta
from typing import Any

from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

# ── Exports ────────────────────────────────────────────────────────────────────
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

app = FastAPI(title="LORANNYS Rule Engine API", version="1.0.0")

# CORS — permite peticiones desde el frontend HTML (cualquier origen por defecto)
_origins_raw = os.environ.get("ALLOWED_ORIGINS", "*")
_origins = [o.strip() for o in _origins_raw.split(",")]
app.add_middleware(
    CORSMiddleware,
    allow_origins=_origins,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ══════════════════════════════════════════════════════════════════════════════
# RULE ENGINE — Extrae compromisos de transcripciones en español
# ══════════════════════════════════════════════════════════════════════════════

COMMITMENT_PATTERNS = [
    (r"\bme comprometo a\b", 0.95, "compromiso explícito"),
    (r"\bqueda(n)? pendiente\b", 0.90, "pendiente explícito"),
    (r"\bvamos a (gestionar|enviar|revisar|coordinar|programar|radicar)\b", 0.85, "vamos a + acción"),
    (r"\bdebemos (enviar|hacer|gestionar|revisar|radicar|coordinar)\b", 0.85, "debemos + acción"),
    (r"\byo me encargo de\b", 0.90, "yo me encargo"),
    (r"\bqued(amos|ó|o) en\b", 0.80, "quedamos en"),
    (r"\bvoy a (enviar|hacer|gestionar|llamar|coordinar|revisar)\b", 0.80, "voy a + acción"),
    (r"\bel (próximo|siguiente) paso es\b", 0.75, "próximo paso"),
    (r"\bhay que (enviar|hacer|gestionar|revisar|radicar)\b", 0.65, "hay que + acción"),
    (r"\ble toca a\b", 0.65, "le toca a"),
    (r"\bnos toca\b", 0.65, "nos toca"),
]

UNCERTAINTY_PATTERNS = [r"\bsi\b", r"\btal vez\b", r"\bposiblemente\b", r"\bno creo\b", r"\bdependiendo\b"]

DEADLINE_PATTERNS = [
    (r"\bpara el (lunes|martes|miércoles|jueves|viernes|sábado|domingo)\b", "weekday"),
    (r"\ben (dos|tres|cuatro|\d+) (días|semanas|meses)\b", "relative"),
    (r"\bantes de fin de mes\b", "end_of_month"),
    (r"\bla próxima semana\b", "next_week"),
    (r"\beste mes\b", "this_month"),
    (r"\bpara el (\d{1,2}) de (\w+)\b", "specific_day"),
]

WEEKDAYS = {"lunes": 0, "martes": 1, "miércoles": 2, "jueves": 3, "viernes": 4, "sábado": 5, "domingo": 6}
MONTHS   = {"enero":0,"febrero":1,"marzo":2,"abril":3,"mayo":4,"junio":5,"julio":6,"agosto":7,"septiembre":8,"octubre":9,"noviembre":10,"diciembre":11}
NUM_WORDS = {"dos": 2, "tres": 3, "cuatro": 4}

GAZETTEER = {
    "ibagué": "municipio", "soledad": "municipio", "sincelejo": "municipio", "neiva": "municipio",
    "tolima": "departamento", "atlántico": "departamento", "huila": "departamento", "sucre": "departamento",
    "gobernación": "entidad", "alcaldía": "entidad", "concejo municipal": "entidad",
    "secretaría de salud": "entidad", "secretaría de obras": "entidad",
    "ministerio de vivienda": "entidad", "ministerio de agricultura": "entidad",
    "ica": "entidad", "finagro": "entidad", "urt": "entidad", "uariv": "entidad",
    "junta de acción comunal": "entidad", "ese": "entidad",
    "acueducto": "tema", "alcantarillado": "tema", "regalías": "tema",
    "seguridad": "tema", "vías": "tema", "salud": "tema",
}


import re


def _normalize(text: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFKD", text.lower()) if not unicodedata.combining(c))


def _split_sentences(text: str) -> list[str]:
    text = re.sub(r"\s+", " ", text).strip()
    if not text:
        return []
    return [s.strip() for s in re.split(r"(?<=[.!?])\s+", text) if s.strip()]


def _guess_responsible(sentence: str) -> str | None:
    match = re.search(r"\b([A-ZÁÉÍÓÚÑ][a-záéíóúñ]+\s[A-ZÁÉÍÓÚÑ][a-záéíóúñ]+)\b", sentence)
    return match.group(1) if match else None


def _guess_deadline_text(lower: str) -> str | None:
    for pattern, _ in DEADLINE_PATTERNS:
        m = re.search(pattern, lower, re.IGNORECASE)
        if m:
            return m.group(0)
    return None


def _resolve_deadline(text: str, reference_date: date) -> date | None:
    if not text:
        return None
    lower = text.lower()
    m = re.search(r"para el (lunes|martes|miércoles|jueves|viernes|sábado|domingo)", lower)
    if m:
        target = (WEEKDAYS[m.group(1)] + 1) % 7  # convertir a getDay() JS style
        d = reference_date
        # Encontrar el próximo día de la semana
        for i in range(1, 8):
            nd = d + timedelta(days=i)
            if nd.weekday() == WEEKDAYS[m.group(1)]:
                return nd
    m = re.search(r"en (dos|tres|cuatro|\d+) (días|semanas|meses)", lower)
    if m:
        n = NUM_WORDS.get(m.group(1), None) or int(m.group(1))
        unit = m.group(2)
        if unit == "días":
            return reference_date + timedelta(days=n)
        elif unit == "semanas":
            return reference_date + timedelta(weeks=n)
        else:
            return date(reference_date.year, reference_date.month + n if reference_date.month + n <= 12 else 1,
                        reference_date.day)
    if "antes de fin de mes" in lower or "este mes" in lower:
        return date(reference_date.year, reference_date.month,
                    (date(reference_date.year, reference_date.month % 12 + 1, 1) - timedelta(days=1)).day)
    if "la próxima semana" in lower:
        return reference_date + timedelta(weeks=1)
    m = re.search(r"para el (\d{1,2}) de (\w+)", lower)
    if m:
        day = int(m.group(1))
        mon = MONTHS.get(m.group(2))
        if mon is not None:
            d = date(reference_date.year, mon + 1, day)
            if d < reference_date:
                d = date(d.year + 1, d.month, d.day)
            return d
    return None


class AnalyzeRequest(BaseModel):
    text: str
    meeting_date: date | None = None
    known_contacts: list[dict] = []   # [{id, name}] — para matchear personas


@app.get("/health")
def health():
    return {"status": "ok", "service": "lorannys-rule-engine"}


@app.post("/analyze")
def analyze_transcript(payload: AnalyzeRequest):
    """
    Analiza una transcripción en español y extrae compromisos y entidades.
    No persiste nada — solo devuelve el análisis para que el frontend lo revise.
    """
    reference_date = payload.meeting_date or date.today()
    contact_map = {c["name"].lower(): c["id"] for c in payload.known_contacts}
    sentences = _split_sentences(payload.text)
    commitments = []
    entities = []

    for idx, sentence in enumerate(sentences):
        lower = sentence.lower()

        # Buscar patrones de compromiso
        for pattern_re, weight, label in COMMITMENT_PATTERNS:
            if re.search(pattern_re, lower, re.IGNORECASE):
                confidence = weight
                word_count = len(sentence.split())
                if word_count > 40: confidence -= 0.2
                elif word_count > 25: confidence -= 0.1
                if any(re.search(u, lower, re.IGNORECASE) for u in UNCERTAINTY_PATTERNS):
                    confidence -= 0.25
                confidence = round(max(0.05, min(0.99, confidence)), 2)

                dl_text = _guess_deadline_text(lower)
                resolved = _resolve_deadline(dl_text, reference_date) if dl_text else None

                commitments.append({
                    "sentence_index": idx,
                    "text": sentence,
                    "confidence": confidence,
                    "requires_llm": confidence < 0.6,
                    "matched_pattern": label,
                    "responsible": _guess_responsible(sentence),
                    "deadline_text": dl_text,
                    "resolved_due_date": resolved.isoformat() if resolved else None,
                })
                break

        # Extraer entidades del gazetteer
        for term, entity_type in GAZETTEER.items():
            if re.search(r"\b" + re.escape(term) + r"\b", lower, re.IGNORECASE):
                entities.append({"entity_type": entity_type, "value": term, "confidence": 0.95, "matched_contact_id": None})

        # Detectar personas (nombres propios)
        for m in re.finditer(r"\b([A-ZÁÉÍÓÚÑ][a-záéíóúñ]+\s[A-ZÁÉÍÓÚÑ][a-záéíóúñ]+)\b", sentence):
            name = m.group(1)
            known_id = contact_map.get(name.lower())
            entities.append({
                "entity_type": "persona",
                "value": name,
                "confidence": 1.0 if known_id else 0.4,
                "matched_contact_id": known_id,
            })

    reasons = [
        *[f"Ambiguo ({c['confidence']}): {c['text'][:50]}..." for c in commitments if c["requires_llm"]],
        *([f"Personas sin match: {', '.join({e['value'] for e in entities if e['entity_type']=='persona' and not e['matched_contact_id']})}"]
          if any(e["entity_type"] == "persona" and not e["matched_contact_id"] for e in entities) else []),
    ]

    return {
        "sentences_analyzed": len(sentences),
        "commitments": commitments,
        "entities": entities,
        "llm_escalation_needed": len(reasons) > 0,
        "llm_escalation_reasons": reasons,
    }


# ══════════════════════════════════════════════════════════════════════════════
# EXPORTACIÓN — recibe datos desde el frontend, devuelve el archivo
# ══════════════════════════════════════════════════════════════════════════════

BRAND_PURPLE = "7C3AED"
PLUM = "2B0A3D"
LIGHT_PURPLE = "F5F3FF"


def _xl_sheet(wb, title: str, headers: list[str], rows: list[list], col_widths: list[int]):
    ws = wb.active if wb.active.title == "Sheet" else wb.create_sheet(title)
    ws.title = title
    ws.row_dimensions[1].height = 22
    fill = PatternFill("solid", fgColor=BRAND_PURPLE)
    font_h = Font(bold=True, color="FFFFFF", size=11)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = fill; cell.font = font_h
        cell.alignment = Alignment(horizontal="left", vertical="center")
    alt_fill = PatternFill("solid", fgColor=LIGHT_PURPLE)
    for r_idx, row in enumerate(rows, 2):
        bg = alt_fill if r_idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=col, value=val)
            cell.fill = bg
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w
    return ws


def _pdf_style(col_widths):
    from reportlab.lib.colors import HexColor
    return TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), HexColor(f"#{BRAND_PURPLE}")),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, 0), 9),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
        ("FONTSIZE",   (0, 1), (-1, -1), 8),
        ("TOPPADDING", (0, 1), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 5),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, HexColor(f"#{LIGHT_PURPLE}")]),
        ("GRID", (0, 0), (-1, -1), 0.3, HexColor("#E5E7EB")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ])


class ExportRequest(BaseModel):
    data: list[dict[str, Any]]


def _stream_xlsx(wb: Workbook, filename: str) -> StreamingResponse:
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename={filename}"})


def _stream_pdf(story: list, filename: str) -> StreamingResponse:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=2*cm, bottomMargin=1.5*cm)
    doc.build(story); buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": f"attachment; filename={filename}"})


@app.post("/export/contacts.xlsx")
def export_contacts_xlsx(req: ExportRequest):
    wb = Workbook()
    _xl_sheet(wb, "Contactos",
        ["Nombre", "Cargo", "Organización", "Departamento", "Municipio", "Categorías", "Empresa", "Score"],
        [[c.get("name",""), c.get("cargo",""), c.get("organizacion",""), c.get("departamento",""),
          c.get("municipio",""), c.get("categorias",""), c.get("company_name",""), c.get("score","")]
         for c in req.data],
        [28, 30, 30, 18, 18, 30, 30, 8])
    return _stream_xlsx(wb, "contactos.xlsx")


@app.post("/export/contacts.pdf")
def export_contacts_pdf(req: ExportRequest):
    from reportlab.lib.colors import HexColor
    small = ParagraphStyle("cell", fontName="Helvetica", fontSize=7.5, leading=10)
    title_style = ParagraphStyle("t", fontName="Helvetica-Bold", fontSize=16, textColor=HexColor(f"#{PLUM}"), spaceAfter=4)
    rows = [["Nombre", "Cargo / Organización", "Territorio", "Categorías", "Score"]]
    for c in req.data:
        territorio = ", ".join(filter(None, [c.get("municipio"), c.get("departamento")])) or "—"
        rows.append([
            Paragraph(c.get("name",""), small),
            Paragraph(f"{c.get('cargo','')} / {c.get('organizacion','')}".strip(" /"), small),
            Paragraph(territorio, small),
            Paragraph(c.get("categorias","—"), small),
            Paragraph(str(round((c.get("score") or 0)*100)), small),
        ])
    col_widths = [5*cm, 7*cm, 5*cm, 5*cm, 1.8*cm]
    story = [
        Paragraph("Directorio de Contactos", title_style),
        Paragraph(f"Generado: {date.today().strftime('%d de %B de %Y')} · {len(req.data)} contacto(s)",
                  ParagraphStyle("sub", fontName="Helvetica", fontSize=9, textColor=colors.grey, spaceAfter=14)),
        Table(rows, colWidths=col_widths, repeatRows=1),
    ]
    story[-1].setStyle(_pdf_style(col_widths))
    return _stream_pdf(story, "contactos.pdf")


@app.post("/export/commitments.xlsx")
def export_commitments_xlsx(req: ExportRequest):
    wb = Workbook()
    _xl_sheet(wb, "Compromisos",
        ["Compromiso", "Contacto", "Responsable", "Estado", "Fecha límite", "Creado"],
        [[c.get("text","")[:100], c.get("contact_name","—"), c.get("responsible","—"),
          c.get("status",""), str(c.get("due_date") or "—"), str(c.get("created_at",""))[:10]]
         for c in req.data],
        [60, 25, 25, 18, 15, 13])
    return _stream_xlsx(wb, "compromisos.xlsx")


@app.post("/export/commitments.pdf")
def export_commitments_pdf(req: ExportRequest):
    from reportlab.lib.colors import HexColor
    small = ParagraphStyle("cell", fontName="Helvetica", fontSize=7.5, leading=10)
    title_style = ParagraphStyle("t", fontName="Helvetica-Bold", fontSize=16, textColor=HexColor(f"#{PLUM}"), spaceAfter=4)
    rows = [["Compromiso", "Contacto", "Responsable", "Estado", "Vencimiento"]]
    for c in req.data:
        rows.append([
            Paragraph((c.get("text",""))[:100], small),
            Paragraph(c.get("contact_name","—"), small),
            Paragraph(c.get("responsible","—"), small),
            Paragraph(c.get("status",""), small),
            Paragraph(str(c.get("due_date") or "—"), small),
        ])
    col_widths = [9*cm, 4.5*cm, 4*cm, 2.5*cm, 2.5*cm]
    story = [
        Paragraph("Compromisos del Despacho", title_style),
        Paragraph(f"Generado: {date.today().strftime('%d de %B de %Y')} · {len(req.data)} compromiso(s)",
                  ParagraphStyle("sub", fontName="Helvetica", fontSize=9, textColor=colors.grey, spaceAfter=14)),
        Table(rows, colWidths=col_widths, repeatRows=1),
    ]
    story[-1].setStyle(_pdf_style(col_widths))
    return _stream_pdf(story, "compromisos.pdf")


@app.post("/export/meetings.xlsx")
def export_meetings_xlsx(req: ExportRequest):
    wb = Workbook()
    _xl_sheet(wb, "Reuniones",
        ["Fecha reunión", "Contacto", "Oraciones analizadas", "Compromisos", "Extracto"],
        [[str(m.get("meeting_date") or m.get("created_at",""))[:10], m.get("contact_name","—"),
          m.get("sentences_analyzed",0), m.get("commitments_count",0),
          (m.get("transcript_text",""))[:200]]
         for m in req.data],
        [14, 26, 12, 12, 80])
    return _stream_xlsx(wb, "reuniones.xlsx")
